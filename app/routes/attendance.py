import io
from datetime import datetime, timedelta
from typing import Dict, List

from flask import Blueprint, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..models.user import User
from ..models.subject import Subject
from ..models.attendance import Attendance
from .. import db_service, ai_service

bp = Blueprint('attendance', __name__, url_prefix='/attendance')

@bp.route('', methods=['GET'])
@login_required
def get_attendance():
    """Get attendance records with filters"""
    try:
        # Parse query parameters
        student_id = request.args.get('student_id')
        subject_id = request.args.get('subject_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        # Validate date formats
        if start_date:
            try:
                datetime.strptime(start_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400
        if end_date:
            try:
                # Add time to include the entire day
                dt = datetime.strptime(end_date, '%Y-%m-%d')
                end_date = (dt + timedelta(days=1)).isoformat()
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400

        # Get records based on user role
        if current_user.is_student:
            records = Attendance.get_by_student(
                student_id=current_user.id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date
            )
        elif current_user.is_teacher:
            if subject_id and subject_id not in current_user.classes:
                return jsonify({'error': 'Unauthorized to view this subject'}), 403
            records = db_service.get_attendance(
                student_id=student_id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date,
                teacher_classes=current_user.classes
            )
        else:  # Admin
            records = db_service.get_attendance(
                student_id=student_id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date
            )

        # Paginate results
        total_records = len(records)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_records = records[start_idx:end_idx]

        # Get AI insights if requested
        insights = None
        if request.args.get('include_insights') == 'true':
            insights = ai_service.analyze_attendance_patterns(records)

        return jsonify({
            'records': [record.to_dict() for record in paginated_records],
            'total_records': total_records,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_records + per_page - 1) // per_page,
            'insights': insights.get('analysis') if insights else None
        })

    except Exception as e:
        current_app.logger.error(f"Error getting attendance: {str(e)}")
        return jsonify({'error': str(e)}), 500

@bp.route('/download', methods=['GET'])
@login_required
def download_attendance():
    """Download attendance records as Excel file"""
    try:
        # Get filter parameters
        student_id = request.args.get('student_id')
        subject_id = request.args.get('subject_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        report_type = request.args.get('type', 'detailed')  # simple or detailed

        # Get records based on user role
        if current_user.is_student:
            records = Attendance.get_by_student(
                student_id=current_user.id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date
            )
        elif current_user.is_teacher:
            if subject_id and subject_id not in current_user.classes:
                return jsonify({'error': 'Unauthorized to download this subject'}), 403
            records = db_service.get_attendance(
                student_id=student_id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date,
                teacher_classes=current_user.classes
            )
        else:  # Admin
            records = db_service.get_attendance(
                student_id=student_id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date
            )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        centered = Alignment(horizontal="center")

        # Write headers
        if report_type == 'detailed':
            headers = [
                'Date', 'Time', 'Student ID', 'Student Name', 'Subject Code',
                'Subject Name', 'Status', 'Verification Method', 'Confidence',
                'Marked By', 'Notes'
            ]
        else:
            headers = ['Date', 'Student ID', 'Student Name', 'Subject', 'Status']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered

        # Write data
        for row, record in enumerate(records, 2):
            if report_type == 'detailed':
                ws.cell(row=row, column=1, value=record.date)
                ws.cell(row=row, column=2, value=record.time)
                ws.cell(row=row, column=3, value=record.student_id)
                ws.cell(row=row, column=4, value=record.student_name)
                ws.cell(row=row, column=5, value=record.subject_id)
                ws.cell(row=row, column=6, value=record.subject_name)
                ws.cell(row=row, column=7, value=record.status)
                ws.cell(row=row, column=8, value='Face Recognition' if record.face_id else 'Manual')
                ws.cell(row=row, column=9, value=f"{record.confidence:.2f}%" if record.confidence else 'N/A')
                ws.cell(row=row, column=10, value=record.marked_by)
                ws.cell(row=row, column=11, value=record.notes)
            else:
                ws.cell(row=row, column=1, value=record.date)
                ws.cell(row=row, column=2, value=record.student_id)
                ws.cell(row=row, column=3, value=record.student_name)
                ws.cell(row=row, column=4, value=record.subject_name)
                ws.cell(row=row, column=5, value=record.status)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'attendance_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        current_app.logger.error(f"Error downloading attendance: {str(e)}")
        return jsonify({'error': str(e)}), 500

@bp.route('/upload', methods=['POST'])
@login_required
def upload_attendance():
    """Upload attendance records from Excel file"""
    try:
        if not current_user.is_teacher and not current_user.is_admin:
            return jsonify({'error': 'Unauthorized'}), 403

        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400

        # Load workbook
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Get headers
        headers = [cell.value for cell in ws[1]]
        required_headers = ['Student ID', 'Student Name', 'Subject Code', 'Status']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({
                'error': f"Missing required columns: {', '.join(missing_headers)}"
            }), 400

        # Process records
        records = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            try:
                # Get values
                data = {headers[i]: cell.value for i, cell in enumerate(row)}
                
                # Validate student
                student = User.get_by_id(data['Student ID'])
                if not student or not student.is_student:
                    errors.append(f"Row {row_idx}: Invalid student ID")
                    continue

                # Validate subject
                subject = Subject.get_by_code(data['Subject Code'])
                if not subject:
                    errors.append(f"Row {row_idx}: Invalid subject code")
                    continue

                # Verify teacher has access to subject
                if current_user.is_teacher and subject.id not in current_user.classes:
                    errors.append(f"Row {row_idx}: Unauthorized to mark attendance for this subject")
                    continue

                # Create attendance record
                attendance = Attendance.create(
                    student_id=student.id,
                    student_name=student.name,
                    subject_id=subject.id,
                    subject_name=subject.name,
                    status=data['Status'],
                    marked_by=current_user.id,
                    notes=data.get('Notes')
                )
                if attendance:
                    records.append(attendance.to_dict())

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        return jsonify({
            'message': f"Processed {len(records)} records with {len(errors)} errors",
            'records': records,
            'errors': errors
        })

    except Exception as e:
        current_app.logger.error(f"Error uploading attendance: {str(e)}")
        return jsonify({'error': str(e)}), 500

@bp.route('/report', methods=['GET'])
@login_required
def generate_report():
    """Generate attendance report with AI insights"""
    try:
        # Get parameters
        subject_id = request.args.get('subject_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        report_type = request.args.get('type', 'summary')  # summary, detailed, trends

        # Validate subject access
        if subject_id:
            subject = Subject.get_by_id(subject_id)
            if not subject:
                return jsonify({'error': 'Subject not found'}), 404
            if current_user.is_teacher and subject.id not in current_user.classes:
                return jsonify({'error': 'Unauthorized to view this subject'}), 403

        # Get attendance records
        if current_user.is_student:
            records = Attendance.get_by_student(
                student_id=current_user.id,
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date
            )
        elif current_user.is_teacher:
            records = db_service.get_attendance(
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date,
                teacher_classes=current_user.classes
            )
        else:  # Admin
            records = db_service.get_attendance(
                subject_id=subject_id,
                start_date=start_date,
                end_date=end_date
            )

        # Generate AI report
        report = ai_service.generate_attendance_report(
            [r.to_dict() for r in records],
            report_type=report_type
        )

        return jsonify(report)

    except Exception as e:
        current_app.logger.error(f"Error generating report: {str(e)}")
        return jsonify({'error': str(e)}), 500 