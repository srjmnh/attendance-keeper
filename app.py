import os
import sys
import base64
import json
import io
from datetime import datetime
import logging

import cv2
import numpy as np
from PIL import Image, ImageEnhance
from flask import (
    Flask,
    request,
    jsonify,
    render_template,
    render_template_string,
    send_file,
    redirect,
    url_for,
    flash,
    session
)
from flask_login import (
    LoginManager,
    login_user,
    login_required,
    logout_user,
    current_user,
    UserMixin,
)
from werkzeug.security import generate_password_hash, check_password_hash

# -----------------------------
# 1) AWS Rekognition Setup
# -----------------------------
import boto3

AWS_ACCESS_KEY_ID = os.getenv('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.getenv('AWS_SECRET_ACCESS_KEY')
AWS_REGION = os.getenv('AWS_REGION', 'us-east-1')
COLLECTION_ID = "students"

rekognition_client = boto3.client(
    'rekognition',
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    region_name=AWS_REGION
)

def create_collection_if_not_exists(collection_id):
    try:
        rekognition_client.create_collection(CollectionId=collection_id)
        print(f"Collection '{collection_id}' created.")
    except rekognition_client.exceptions.ResourceAlreadyExistsException:
        print(f"Collection '{collection_id}' already exists.")

create_collection_if_not_exists(COLLECTION_ID)

# -----------------------------
# 2) Firebase Firestore Setup
# -----------------------------
import firebase_admin
from firebase_admin import credentials, firestore

base64_cred_str = os.environ.get("FIREBASE_ADMIN_CREDENTIALS_BASE64")
if not base64_cred_str:
    raise ValueError("FIREBASE_ADMIN_CREDENTIALS_BASE64 not found in environment.")

decoded_cred_json = base64.b64decode(base64_cred_str)
cred_dict = json.loads(decoded_cred_json)
cred = credentials.Certificate(cred_dict)
firebase_admin.initialize_app(cred)
db = firestore.client()

# -----------------------------
# 3) Gemini Chatbot Setup
# -----------------------------
import google.generativeai as genai

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY environment variable not set.")

genai.configure(api_key=GEMINI_API_KEY)
# If you do NOT have access to "gemini-1.5-flash", switch to "models/chat-bison-001"
model = genai.GenerativeModel("models/gemini-1.5-flash")

# Chat memory
MAX_MEMORY = 20
conversation_memory = []

# A big system prompt describing the entire system
system_context = """You are Gemini, a somewhat witty (but polite) AI assistant.
Facial Recognition Attendance system features:

1) AWS Rekognition:
   - We keep a 'students' collection on startup.
   - /register indexes a face by (name + student_id).
   - /recognize detects faces in an uploaded image, logs attendance in Firestore if matched.

2) Attendance:
   - Firestore 'attendance' collection: { student_id, name, timestamp, subject_id, subject_name, status='PRESENT' }.
   - UI has tabs: Register, Recognize, Subjects, Attendance (Bootstrap + DataTables).
   - Attendance can filter, inline-edit, download/upload Excel.

3) Subjects:
   - We can add subjects to 'subjects' collection, referenced in recognition.

4) Multi-Face:
   - If multiple recognized faces, each is logged to attendance.

5) Chat:
   - You are the assistant, a bit humorous, guiding usage or code features.
"""

# Start the conversation with a system message
conversation_memory.append({"role": "system", "content": system_context})

# -----------------------------
# 4) Flask App Initialization
# -----------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "your_default_secret_key")

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # Redirect to 'login' when login is required

# -----------------------------
# 5) User Model and Authentication
# -----------------------------

class User(UserMixin):
    def __init__(self, id, username, password_hash, role, classes=None):
        self.id = id
        self.username = username
        self.password_hash = password_hash
        self.role = role  # 'admin', 'teacher', 'student'
        self.classes = classes or []  # List of class IDs (for teachers)

# User loader callback for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    user_doc = db.collection("users").document(user_id).get()
    if user_doc.exists:
        user_data = user_doc.to_dict()
        return User(
            id=user_doc.id,
            username=user_data.get("username"),
            password_hash=user_data.get("password_hash"),
            role=user_data.get("role"),
            classes=user_data.get("classes", [])
        )
    return None

# Define role_required decorator
from functools import wraps

def role_required(required_roles):
    def decorator(f):
        @wraps(f)
        @login_required
        def wrapped(*args, **kwargs):
            if current_user.role not in required_roles:
                flash("You do not have permission to access this page.", "danger")
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return wrapped
    return decorator

# -----------------------------
# 6) Login and Registration Routes
# -----------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username").strip()
        password = request.form.get("password").strip()
        if not username or not password:
            flash("Please enter both username and password.", "warning")
            return render_template("login.html")
        
        # Query Firestore for user
        users_ref = db.collection("users")
        query = users_ref.where("username", "==", username).stream()
        user = None
        for doc in query:
            user_data = doc.to_dict()
            if check_password_hash(user_data.get("password_hash", ""), password):
                user = User(
                    id=doc.id,
                    username=user_data.get("username"),
                    password_hash=user_data.get("password_hash"),
                    role=user_data.get("role"),
                    classes=user_data.get("classes", [])
                )
                break

        if user:
            login_user(user)
            flash("Logged in successfully!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "danger")
            return render_template("login.html")
    else:
        return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))

# -----------------------------
# 7) Admin Routes for User Management
# -----------------------------

@app.route("/admin/manage_users", methods=["GET", "POST"])
@role_required(['admin'])
def admin_panel():
    if request.method == "POST":
        # Handle user creation logic
        # Ensure role is assigned
        username = request.form.get("username").strip()
        password = request.form.get("password").strip()
        role = request.form.get("role").strip()

        if not username or not password or not role:
            flash("All fields are required.", "warning")
            return redirect(url_for('admin_panel'))

        # Validate role
        if role not in ['admin', 'teacher', 'student']:
            flash("Invalid role selected.", "danger")
            return redirect(url_for('admin_panel'))

        # Hash the password
        password_hash = generate_password_hash(password, method="pbkdf2:sha256")

        # Create user in Firestore
        try:
            db.collection("users").add({
                "username": username,
                "password_hash": password_hash,
                "role": role,
                "classes": []  # Initialize with empty classes
            })
            flash(f"User '{username}' with role '{role}' created successfully!", "success")
        except Exception as e:
            flash(f"Error creating user: {str(e)}", "danger")

        return redirect(url_for('admin_panel'))

    # GET request - display users
    users = db.collection("users").stream()
    users_list = []
    for user in users:
        user_data = user.to_dict()
        users_list.append({
            "username": user_data.get("username", "N/A"),
            "role": user_data.get("role") if user_data.get("role") else "N/A",
            "classes": user_data.get("classes", [])
        })

    return render_template("admin_panel.html", users=users_list)

@app.route("/admin/create_user", methods=["POST"])
@role_required(['admin'])
def create_user():
    username = request.form.get("username").strip()
    password = request.form.get("password").strip()
    role = request.form.get("role").strip()

    if not username or not password or not role:
        flash("All fields are required.", "warning")
        return redirect(url_for('admin_panel'))

    # Validate role
    if role not in ['admin', 'teacher', 'student']:
        flash("Invalid role selected.", "danger")
        return redirect(url_for('admin_panel'))

    # Hash the password
    password_hash = generate_password_hash(password, method="pbkdf2:sha256")

    # Create user in Firestore
    try:
        db.collection("users").add({
            "username": username,
            "password_hash": password_hash,
            "role": role,
            "classes": []  # Initialize with empty classes
        })
        flash(f"User '{username}' with role '{role}' created successfully!", "success")
    except Exception as e:
        flash(f"Error creating user: {str(e)}", "danger")

    return redirect(url_for('admin_panel'))

# -----------------------------
# 8) Single-Page HTML + Chat Widget with Role-Based Tabs
# -----------------------------

@app.route("/")
@app.route("/dashboard")
@login_required
def dashboard():
    # Get subjects for the dropdown
    subjects = []
    if current_user.role in ['admin', 'teacher']:
        subjects_ref = db.collection("subjects").stream()
        for doc in subjects_ref:
            subjects.append({
                'id': doc.id,
                'name': doc.to_dict().get('name', '')
            })
    
    # Get users list for admin panel
    users = []
    if current_user.role == 'admin':
        users_ref = db.collection("users").stream()
        for doc in users_ref:
            user_data = doc.to_dict()
            users.append({
                'username': user_data.get('username', ''),
                'role': user_data.get('role', ''),
                'classes': user_data.get('classes', [])
            })
    
    # Determine which tab is active based on query parameter
    active_tab = request.args.get("tab", "recognize")
    
    return render_template(
        "dashboard.html",
        active_tab=active_tab,
        role=current_user.role,
        subjects=subjects,
        users=users
    )

# -----------------------------
# 9) Routes (Register, Recognize, Subjects, Attendance)
# -----------------------------

# Register Face (GET/POST)
@app.route("/register", methods=["GET","POST"])
@role_required(['admin', 'teacher'])
def register_face():
    if request.method == "GET":
        return "Welcome to /register. Please POST with {name, student_id, image} to register."

    data = request.json
    name = data.get('name')
    student_id = data.get('student_id')
    image = data.get('image')
    if not name or not student_id or not image:
        return jsonify({"message": "Missing name, student_id, or image"}), 400

    sanitized_name = "".join(c if c.isalnum() or c in "_-." else "_" for c in name)
    image_data = image.split(",")[1]
    image_bytes = base64.b64decode(image_data)

    # Enhance image before indexing
    pil_image = Image.open(io.BytesIO(image_bytes))
    enhanced_image = enhance_image(pil_image)
    buffered = io.BytesIO()
    enhanced_image.save(buffered, format="JPEG")
    enhanced_image_bytes = buffered.getvalue()

    external_image_id = f"{sanitized_name}_{student_id}"
    try:
        response = rekognition_client.index_faces(
            CollectionId=COLLECTION_ID,
            Image={'Bytes': enhanced_image_bytes},
            ExternalImageId=external_image_id,
            DetectionAttributes=['ALL'],
            QualityFilter='AUTO'
        )
    except Exception as e:
        return jsonify({"message": f"Failed to index face: {str(e)}"}), 500

    if not response.get('FaceRecords'):
        return jsonify({"message": "No face detected in the image"}), 400

    return jsonify({"message": f"Student {name} with ID {student_id} registered successfully!"}), 200

# Recognize Face (GET/POST)
@app.route("/recognize", methods=["GET","POST"])
@login_required
def recognize_face():
    if request.method == "GET":
        return "Welcome to /recognize. Please POST with {image, subject_id(optional)} to detect faces."

    data = request.json
    image_str = data.get('image')
    subject_id = data.get('subject_id') or ""
    if not image_str:
        return jsonify({"message": "No image provided"}), 400

    # Optionally fetch subject name
    subject_name = ""
    if subject_id:
        sdoc = db.collection("subjects").document(subject_id).get()
        if sdoc.exists:
            subject_name = sdoc.to_dict().get("name", "")
        else:
            subject_name = "Unknown Subject"

    image_data = image_str.split(",")[1]
    image_bytes = base64.b64decode(image_data)

    # Enhance image before detection
    pil_image = Image.open(io.BytesIO(image_bytes))
    enhanced_image = enhance_image(pil_image)
    buffered = io.BytesIO()
    enhanced_image.save(buffered, format="JPEG")
    enhanced_image_bytes = buffered.getvalue()

    try:
        # Detect faces in the image
        detect_response = rekognition_client.detect_faces(
            Image={'Bytes': enhanced_image_bytes},
            Attributes=['ALL']
        )
    except Exception as e:
        return jsonify({"message": f"Failed to detect faces: {str(e)}"}), 500

    faces = detect_response.get('FaceDetails', [])
    face_count = len(faces)
    identified_people = []

    if face_count == 0:
        return jsonify({
            "message": "No faces detected in the image.",
            "total_faces": face_count,
            "identified_people": identified_people
        }), 200

    for idx, face in enumerate(faces):
        # Rekognition provides bounding box coordinates relative to image dimensions
        bbox = face['BoundingBox']
        pil_img = Image.open(io.BytesIO(enhanced_image_bytes))
        img_width, img_height = pil_img.size

        left = int(bbox['Left'] * img_width)
        top = int(bbox['Top'] * img_height)
        width = int(bbox['Width'] * img_width)
        height = int(bbox['Height'] * img_height)
        right = left + width
        bottom = top + height

        # Crop the face from the image
        cropped_face = pil_img.crop((left, top, right, bottom))
        buffer = io.BytesIO()
        cropped_face.save(buffer, format="JPEG")
        cropped_face_bytes = buffer.getvalue()

        try:
            # Search for the face in the collection
            search_response = rekognition_client.search_faces_by_image(
                CollectionId=COLLECTION_ID,
                Image={'Bytes': cropped_face_bytes},
                MaxFaces=1,
                FaceMatchThreshold=60
            )
        except Exception as e:
            identified_people.append({
                "message": f"Error searching face {idx+1}: {str(e)}",
                "confidence": "N/A"
            })
            continue

        matches = search_response.get('FaceMatches', [])
        if not matches:
            identified_people.append({
                "message": "Face not recognized",
                "confidence": "N/A"
            })
            continue

        match = matches[0]
        ext_id = match['Face']['ExternalImageId']
        confidence = match['Face']['Confidence']

        parts = ext_id.split("_", 1)
        if len(parts) == 2:
            rec_name, rec_id = parts
        else:
            rec_name, rec_id = ext_id, "Unknown"

        identified_people.append({
            "name": rec_name,
            "student_id": rec_id,
            "confidence": confidence
        })

        # If recognized, log attendance
        if rec_id != "Unknown":
            doc = {
                "student_id": rec_id,
                "name": rec_name,
                "timestamp": datetime.utcnow().isoformat(),
                "subject_id": subject_id,
                "subject_name": subject_name,
                "status": "PRESENT"
            }
            db.collection("attendance").add(doc)

    return jsonify({
        "message": f"{face_count} face(s) detected in the photo.",
        "total_faces": face_count,
        "identified_people": identified_people
    }), 200

# SUBJECTS
@app.route("/add_subject", methods=["POST"])
@role_required(['admin', 'teacher'])
def add_subject():
    data = request.json
    subject_name = data.get("subject_name")
    if not subject_name:
        return jsonify({"error": "No subject_name provided"}), 400
    doc_ref = db.collection("subjects").document()
    doc_ref.set({
        "name": subject_name.strip(),
        "created_at": datetime.utcnow().isoformat()
    })
    return jsonify({"message": f"Subject '{subject_name}' added successfully!"}), 200

@app.route("/get_subjects", methods=["GET"])
@login_required
def get_subjects():
    if current_user.role == 'teacher':
        # Teachers can see only subjects they teach
        subjects = []
        for cls in current_user.classes:
            sub_doc = db.collection("subjects").where("name", "==", cls).stream()
            for doc in sub_doc:
                subjects.append({"id": doc.id, "name": doc.to_dict().get("name","")})
        return jsonify({"subjects": subjects}), 200
    else:
        # Admin can see all subjects
        subs = db.collection("subjects").stream()
        subj_list = []
        for s in subs:
            d = s.to_dict()
            subj_list.append({"id": s.id, "name": d.get("name","")})
        return jsonify({"subjects": subj_list}), 200

# ATTENDANCE
import openpyxl
from openpyxl import Workbook

@app.route("/api/attendance", methods=["GET"])
@login_required
def get_attendance():
    student_id = request.args.get("student_id")
    subject_id = request.args.get("subject_id")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = db.collection("attendance")
    if current_user.role == 'teacher':
        # Teachers can view only their classes
        query = query.where("subject_id", "in", current_user.classes)

    if student_id:
        query = query.where("student_id", "==", student_id)
    if subject_id:
        query = query.where("subject_id", "==", subject_id)
    if start_date:
        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.where("timestamp", ">=", dt_start.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid start_date format. Use YYYY-MM-DD."}), 400
    if end_date:
        try:
            dt_end = datetime.strptime(end_date, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, microsecond=999999
            )
            query = query.where("timestamp", "<=", dt_end.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid end_date format. Use YYYY-MM-DD."}), 400

    # Execute query
    try:
        results = query.stream()
    except Exception as e:
        return jsonify({"error": f"Failed to retrieve attendance: {str(e)}"}), 500

    out_list = []
    for doc_ in results:
        dd = doc_.to_dict()
        dd["doc_id"] = doc_.id
        out_list.append(dd)

    return jsonify(out_list)

@app.route("/api/attendance/update", methods=["POST"])
@role_required(['admin', 'teacher'])
def update_attendance():
    data = request.json
    records = data.get("records", [])
    for rec in records:
        doc_id = rec.get("doc_id")
        if not doc_id:
            continue
        # Teachers can update only their classes
        record_doc = db.collection("attendance").document(doc_id).get()
        if not record_doc.exists:
            continue
        record_data = record_doc.to_dict()
        if current_user.role == 'teacher' and record_data.get("subject_id") not in current_user.classes:
            continue  # Skip if subject not assigned to teacher
        ref = db.collection("attendance").document(doc_id)
        update_data = {
            "student_id": rec.get("student_id",""),
            "name": rec.get("name",""),
            "subject_id": rec.get("subject_id",""),
            "subject_name": rec.get("subject_name",""),
            "timestamp": rec.get("timestamp",""),
            "status": rec.get("status","")
        }
        try:
            ref.update(update_data)
        except Exception as e:
            continue  # Optionally, handle errors
    return jsonify({"message": "Attendance records updated successfully."})

@app.route("/api/attendance/download", methods=["GET"])
@login_required
def download_attendance_excel():
    student_id = request.args.get("student_id")
    subject_id = request.args.get("subject_id")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = db.collection("attendance")
    if current_user.role == 'teacher':
        # Teachers can download only their classes
        query = query.where("subject_id", "in", current_user.classes)

    if student_id:
        query = query.where("student_id", "==", student_id)
    if subject_id:
        query = query.where("subject_id", "==", subject_id)
    if start_date:
        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.where("timestamp", ">=", dt_start.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid start_date format. Use YYYY-MM-DD."}), 400
    if end_date:
        try:
            dt_end = datetime.strptime(end_date, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, microsecond=999999
            )
            query = query.where("timestamp", "<=", dt_end.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid end_date format. Use YYYY-MM-DD."}), 400

    # Execute query
    try:
        results = query.stream()
    except Exception as e:
        return jsonify({"error": f"Failed to retrieve attendance: {str(e)}"}), 500

    att_list = []
    for doc_ in results:
        dd = doc_.to_dict()
        dd["doc_id"] = doc_.id
        att_list.append(dd)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["doc_id", "student_id", "name", "subject_id", "subject_name", "timestamp", "status"]
    ws.append(headers)

    for record in att_list:
        row = [
            record.get("doc_id",""),
            record.get("student_id",""),
            record.get("name",""),
            record.get("subject_id",""),
            record.get("subject_name",""),
            record.get("timestamp",""),
            record.get("status","")
        ]
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="attendance.xlsx"
    )

@app.route("/api/attendance/template", methods=["GET"])
@login_required
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    headers = ["doc_id","student_id","name","subject_id","subject_name","timestamp","status"]
    ws.append(headers)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="attendance_template.xlsx"
    )

@app.route("/api/attendance/upload", methods=["POST"])
@role_required(['admin', 'teacher'])
def upload_attendance_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Please upload a .xlsx file"}), 400

    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        return jsonify({"error": f"Failed to read Excel file: {str(e)}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    expected = ("doc_id","student_id","name","subject_id","subject_name","timestamp","status")
    if not rows or rows[0] != expected:
        return jsonify({"error": "Incorrect template format"}), 400

    for row in rows[1:]:
        doc_id, student_id, name, subject_id, subject_name, timestamp, status = row
        if current_user.role == 'teacher' and subject_id not in current_user.classes:
            continue  # Teachers cannot modify other classes' data
        if doc_id:
            doc_data = {
                "student_id": student_id or "",
                "name": name or "",
                "subject_id": subject_id or "",
                "subject_name": subject_name or "",
                "timestamp": timestamp or "",
                "status": status or ""
            }
            try:
                db.collection("attendance").document(doc_id).set(doc_data, merge=True)
            except Exception as e:
                continue  # Optionally, handle errors
        else:
            new_doc = {
                "student_id": student_id or "",
                "name": name or "",
                "subject_id": subject_id or "",
                "subject_name": subject_name or "",
                "timestamp": timestamp or "",
                "status": status or ""
            }
            try:
                db.collection("attendance").add(new_doc)
            except Exception as e:
                continue  # Optionally, handle errors

    return jsonify({"message": "Excel data imported successfully."})

# -----------------------------
# 10) Admin Gemini Chat Integration
# -----------------------------

# Note: Since Gemini is integrated into the main chat widget, and Admin has access to /admin panel,
# you can add additional chatbot functionalities within the Admin dashboard as needed.

# -----------------------------
# 11) Image Enhancement Function (Using OpenCV and Pillow)
# -----------------------------
def enhance_image(pil_image):
    """
    Enhance image quality to improve face detection in distant group photos.
    This includes increasing brightness and contrast.
    """
    # Convert PIL image to OpenCV format
    cv_image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)

    # Increase brightness and contrast
    alpha = 1.2  # Contrast control (1.0-3.0)
    beta = 30    # Brightness control (0-100)
    enhanced_cv_image = cv2.convertScaleAbs(cv_image, alpha=alpha, beta=beta)

    # Convert back to PIL Image
    enhanced_pil_image = Image.fromarray(cv2.cvtColor(enhanced_cv_image, cv2.COLOR_BGR2RGB))

    return enhanced_pil_image

# -----------------------------
# 12) Single-Page HTML + Chat Widget - Protected with Login
# -----------------------------
# Already handled above in INDEX_HTML with role-based tabs

# -----------------------------
# 13) Gemini Chat Endpoint
# -----------------------------
@app.route("/process_prompt", methods=["POST"])
@login_required
@role_required(['admin'])  # Only admin can use Gemini chat
def process_prompt():
    data = request.json
    user_prompt = data.get("prompt","").strip()
    if not user_prompt:
        return jsonify({"error":"No prompt provided"}), 400

    # Add user message
    conversation_memory.append({"role":"user","content":user_prompt})

    # Build conversation string
    conv_str = ""
    for msg in conversation_memory:
        if msg["role"] == "system":
            conv_str += f"System: {msg['content']}\n"
        elif msg["role"] == "user":
            conv_str += f"User: {msg['content']}\n"
        else:
            conv_str += f"Assistant: {msg['content']}\n"

    # Call Gemini
    try:
        response = model.generate_content(conv_str)
    except Exception as e:
        assistant_reply = f"Error generating response: {str(e)}"
    else:
        if not response.candidates:
            assistant_reply = "Hmm, I'm having trouble responding right now."
        else:
            parts = response.candidates[0].content.parts
            assistant_reply = "".join(part.text for part in parts).strip()

    # Add assistant reply
    conversation_memory.append({"role":"assistant","content":assistant_reply})

    if len(conversation_memory) > MAX_MEMORY:
        conversation_memory.pop(0)

    return jsonify({"message": assistant_reply})

# -----------------------------
# 14) Run App
# -----------------------------
def create_default_admin():
    """
    Creates a default admin user if no admin exists in the Firestore 'users' collection.
    """
    admins_ref = db.collection("users").where("role", "==", "admin").stream()
    admins = [admin for admin in admins_ref]

    if not admins:
        default_username = "admin"
        default_password = "Admin123!"  # **Change this password immediately after first login**
        password_hash = generate_password_hash(default_password, method="pbkdf2:sha256")  # Updated method
        
        admin_data = {
            "username": default_username,
            "password_hash": password_hash,
            "role": "admin",
            # 'classes' field is optional for admin
        }
        
        db.collection("users").add(admin_data)
        print(f"Default admin user '{default_username}' created with password '{default_password}'.")
    else:
        print("Admin user already exists. No default admin created.")

@app.route("/change_password", methods=["GET", "POST"])
@role_required(['admin', 'teacher', 'student'])  # Adjust roles as needed
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password")
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")
        
        if not current_password or not new_password or not confirm_password:
            flash("All fields are required.", "danger")
            return redirect(url_for('change_password'))
        
        if new_password != confirm_password:
            flash("New passwords do not match.", "danger")
            return redirect(url_for('change_password'))
        
        # Verify current password
        user_doc = db.collection("users").where("username", "==", current_user.username).stream()
        user = None
        for usr in user_doc:
            user = usr
            break
        
        if user and check_password_hash(user.to_dict().get("password_hash"), current_password):
            # Update password
            new_password_hash = generate_password_hash(new_password, method="pbkdf2:sha256")
            db.collection("users").document(user.id).update({"password_hash": new_password_hash})
            flash("Password updated successfully.", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("Current password is incorrect.", "danger")
            return redirect(url_for('change_password'))
    
    return render_template("change_password.html")

@app.route("/manage_users")
@login_required
@role_required(['admin'])
def manage_users():
    users = db.collection("users").stream()
    user_list = []
    for user in users:
        user_data = user.to_dict()
        user_list.append({
            'username': user_data.get('username', 'N/A'),
            'role': user_data.get('role', 'N/A'),
            # Add other fields as necessary
        })
    return render_template("manage_users.html", users=user_list)

# -----------------------------
# Subjects Management Routes
# -----------------------------

@app.route("/admin/subjects", methods=["GET", "POST"])
@role_required(['admin'])
def manage_subjects():
    if request.method == "POST":
        # Handle adding a new subject
        subject_name = request.form.get("subject_name").strip()
        if not subject_name:
            flash("Subject name cannot be empty.", "warning")
            return redirect(url_for('manage_subjects'))
        
        # Generate a 3-letter subject code
        subject_code = generate_subject_code(subject_name)
        
        # Check if subject_code is unique
        subjects_ref = db.collection("subjects")
        existing = subjects_ref.where("code", "==", subject_code).stream()
        if any(existing):
            flash(f"Subject code '{subject_code}' already exists. Please choose a different subject name.", "danger")
            return redirect(url_for('manage_subjects'))
        
        # Add the new subject to Firestore
        try:
            subjects_ref.add({
                "name": subject_name,
                "code": subject_code.upper(),
                "created_at": datetime.utcnow().isoformat()
            })
            flash(f"Subject '{subject_name}' with code '{subject_code}' added successfully!", "success")
        except Exception as e:
            flash(f"Error adding subject: {str(e)}", "danger")
        
        return redirect(url_for('manage_subjects'))
    
    # GET request - display subjects
    subjects = db.collection("subjects").stream()
    subjects_list = []
    for subject in subjects:
        subject_data = subject.to_dict()
        subjects_list.append({
            'id': subject.id,
            'code': subject_data.get('code', 'N/A'),
            'name': subject_data.get('name', 'N/A'),
            'created_at': subject_data.get('created_at', 'N/A')
        })
    
    return render_template("manage_subjects.html", subjects=subjects_list)

@app.route("/admin/delete_subject/<subject_id>", methods=["POST"])
@login_required
@role_required(['admin'])
def delete_subject(subject_id):
    try:
        db.collection("subjects").document(subject_id).delete()
        flash("Subject deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting subject: {str(e)}", "danger")
    return redirect(url_for('manage_subjects'))

def generate_subject_code(subject_name):
    """
    Generates a 3-letter uppercase subject code based on the subject name.
    Example: "Mathematics" -> "MAT"
    """
    # Take the first letter of up to three words
    letters = [word[0] for word in subject_name.upper().split()[:3]]
    code = ''.join(letters).ljust(3, 'X')  # Pad with 'X' if less than 3 letters
    return code[:3]

@app.route("/admin/update_subject", methods=["POST"])
@role_required(['admin'])
def update_subject():
    data = request.get_json()
    subject_id = data.get("subject_id")
    new_name = data.get("name", "").strip()
    
    if not subject_id or not new_name:
        return jsonify({"error": "Subject ID and new name are required."}), 400
    
    # Generate a new subject code based on the new name
    new_code = generate_subject_code(new_name)
    
    # Check if the new_code is unique
    subjects_ref = db.collection("subjects")
    existing = subjects_ref.where("code", "==", new_code).stream()
    if any(existing):
        return jsonify({"error": f"Subject code '{new_code}' already exists. Please choose a different subject name."}), 400
    
    try:
        subjects_ref.document(subject_id).update({
            "name": new_name,
            "code": new_code
        })
        return jsonify({"message": "Subject updated successfully."}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to update subject: {str(e)}"}), 500

if __name__ == "__main__":
    # Create default admin if none exists
    create_default_admin()
    
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)